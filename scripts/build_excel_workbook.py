#!/usr/bin/env python3
"""
Phase 4G: Build Excel Statistical Analysis Workbook.

Generates an Excel workbook demonstrating advanced Excel skills for A/B test
analysis: Power Query-ready data, statistical formulas, charts, conditional
formatting, and executive dashboard.

Usage:
  python scripts/build_excel_workbook.py

Output:
  output/analysis/experiment_analysis_workbook.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from scripts.excel_data import gather_experiment_data


def _style_header(ws, row: int = 1):
    """Apply header styling."""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")


def build_workbook() -> Path:
    """Create the Excel workbook and return output path."""
    out_dir = ROOT / "output" / "analysis"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "experiment_analysis_workbook.xlsx"

    variant_df, summary_df = gather_experiment_data()
    if variant_df.empty:
        raise RuntimeError("No experiment data. Run analysis assets first.")

    wb = Workbook()
    thin = Side(style="thin")

    # ── Sheet 1: Data Source (for Power Query / CSV import) ─────────────────
    ws1 = wb.active
    ws1.title = "Data_Import"
    ws1["A1"] = "Power Query: Import from output/analysis/experiment_results.csv"
    ws1["A1"].font = Font(bold=True, size=12)
    ws1["A3"] = "Steps:"
    ws1["A4"] = "1. Get Data > From File > From Text/CSV"
    ws1["A5"] = "2. Select: output/analysis/experiment_results.csv"
    ws1["A6"] = "3. Transform: Filter, Add Columns as needed"
    ws1["A7"] = "4. Load to Table"
    ws1["A9"] = "Or run: python scripts/export_experiment_csv.py"
    ws1.column_dimensions["A"].width = 60

    # ── Sheet 2: Experiment Results Table ─────────────────────────────────
    ws2 = wb.create_sheet("Experiment_Results", 1)
    headers = ["Experiment", "Variant", "Users", "Conversions", "Conv_Rate", "Revenue", "AOV"]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2)

    for r, row in variant_df.iterrows():
        ws2.cell(row=r + 2, column=1, value=row["Experiment"])
        ws2.cell(row=r + 2, column=2, value=row["Variant"])
        ws2.cell(row=r + 2, column=3, value=row["Users"])
        ws2.cell(row=r + 2, column=4, value=row["Conversions"])
        ws2.cell(row=r + 2, column=5, value=row["Conv_Rate"] / 100)
        ws2.cell(row=r + 2, column=6, value=row["Revenue"])
        ws2.cell(row=r + 2, column=7, value=row["AOV"])

    n_rows = len(variant_df) + 1
    ws2.column_dimensions["A"].width = 35
    for c in range(2, 8):
        ws2.column_dimensions[get_column_letter(c)].width = 12

    # Add Excel Table
    tab = Table(displayName="ExperimentResults", ref=f"A1:G{n_rows}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False)
    ws2.add_table(tab)

    # Calculated: Lift (relative) - add in col H
    ws2.cell(row=1, column=8, value="Lift_%")
    ws2.cell(row=1, column=8).font = Font(bold=True)
    for r in range(2, n_rows + 1):
        exp = ws2.cell(row=r, column=1).value
        variant = ws2.cell(row=r, column=2).value
        cr = ws2.cell(row=r, column=5).value
        # Find control CR for this experiment
        ctrl_cr = None
        for rr in range(2, n_rows + 1):
            if ws2.cell(row=rr, column=1).value == exp and ws2.cell(row=rr, column=2).value == "Control":
                ctrl_cr = ws2.cell(row=rr, column=5).value
                break
        if ctrl_cr and ctrl_cr > 0 and variant == "Treatment":
            lift = (cr - ctrl_cr) / ctrl_cr * 100
            ws2.cell(row=r, column=8, value=round(lift, 2))
        else:
            ws2.cell(row=r, column=8, value="")

    # ── Sheet 3: Statistical Tests ────────────────────────────────────────
    ws3 = wb.create_sheet("Statistical_Tests", 2)
    ws3["A1"] = "Statistical Tests"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3["A3"] = "Z-Test for Proportions (Conversion Rate)"
    ws3["A4"] = "Formula: =(p_treat-p_ctrl)/SQRT(p_pool*(1-p_pool)*(1/n1+1/n2))"
    ws3["A5"] = "P-Value: =2*(1-NORM.S.DIST(ABS(z_score),TRUE))"
    ws3["A7"] = "Chi-Square Test"
    ws3["A8"] = "Formula: =CHISQ.TEST(actual_range, expected_range)"
    ws3["A10"] = "T-Test for Revenue (Two-Sample)"
    ws3["A11"] = "Formula: =T.TEST(control_revenue, treatment_revenue, 2, 2)"
    ws3["A12"] = "  (2=two-tailed, 2=equal variance)"
    ws3["A14"] = "95% Confidence Interval"
    ws3["A15"] = "Formula: =CONFIDENCE.NORM(0.05, STDEV.S(range), COUNT(range))"
    ws3["A17"] = "Example (Exp 1):"
    ws3.column_dimensions["A"].width = 70

    # Add example values for first experiment
    exp1_ctrl = variant_df[(variant_df["Experiment_ID"] == "exp_001") & (variant_df["Variant"] == "Control")].iloc[0]
    exp1_treat = variant_df[(variant_df["Experiment_ID"] == "exp_001") & (variant_df["Variant"] == "Treatment")].iloc[0]
    n1, n2 = exp1_ctrl["Users"], exp1_treat["Users"]
    c1, c2 = exp1_ctrl["Conversions"], exp1_treat["Conversions"]
    p1, p2 = c1 / n1, c2 / n2
    p_pool = (c1 + c2) / (n1 + n2)
    se = (p_pool * (1 - p_pool) * (1 / n1 + 1 / n2)) ** 0.5
    z = (p2 - p1) / se if se > 0 else 0
    from scipy import stats as sp_stats
    p_val = 2 * (1 - sp_stats.norm.cdf(abs(z)))

    ws3["B19"] = "Control"
    ws3["C19"] = "Treatment"
    ws3["A20"] = "Conversions"
    ws3["B20"] = c1
    ws3["C20"] = c2
    ws3["A21"] = "Users"
    ws3["B21"] = n1
    ws3["C21"] = n2
    ws3["A22"] = "Conversion Rate"
    ws3["B22"] = p1
    ws3["C22"] = p2
    ws3["A23"] = "Z-Score"
    ws3["B23"] = round(z, 4)
    ws3["A24"] = "P-Value (Z)"
    ws3["B24"] = round(p_val, 6)
    ws3["A25"] = "Significant?"
    ws3["B25"] = "Yes" if p_val < 0.05 else "No"

    # ── Sheet 4: Visualizations ───────────────────────────────────────────
    ws4 = wb.create_sheet("Charts", 3)
    ws4["A1"] = "Conversion Rate Comparison (Exp 1)"
    ws4["A1"].font = Font(bold=True, size=12)

    # Data for chart
    chart_data = variant_df[variant_df["Experiment_ID"] == "exp_001"][["Variant", "Conv_Rate"]]
    for r, row in chart_data.iterrows():
        ws4.cell(row=r - chart_data.index[0] + 3, column=1, value=row["Variant"])
        ws4.cell(row=r - chart_data.index[0] + 3, column=2, value=row["Conv_Rate"] / 100)
    ws4["A2"] = "Variant"
    ws4["B2"] = "Conv_Rate"
    _style_header(ws4, 2)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Conversion Rate by Variant"
    chart.y_axis.title = "Conversion Rate"
    chart.x_axis.title = "Variant"
    data = Reference(ws4, min_col=2, min_row=2, max_row=4)
    cats = Reference(ws4, min_col=1, min_row=3, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 18
    ws4.add_chart(chart, "D2")

    # Lift chart data
    ws4["A15"] = "Lift by Experiment"
    ws4["A15"].font = Font(bold=True, size=12)
    lift_rows = []
    for exp_id in variant_df["Experiment_ID"].unique():
        ctrl = variant_df[(variant_df["Experiment_ID"] == exp_id) & (variant_df["Variant"] == "Control")]
        treat = variant_df[(variant_df["Experiment_ID"] == exp_id) & (variant_df["Variant"] == "Treatment")]
        if len(ctrl) and len(treat):
            cr_c, cr_t = ctrl["Conv_Rate"].iloc[0], treat["Conv_Rate"].iloc[0]
            lift = (cr_t - cr_c) / cr_c * 100 if cr_c > 0 else 0
            name = ctrl["Experiment"].iloc[0]
            lift_rows.append((name[:25], round(lift, 1)))
    for i, (name, lift) in enumerate(lift_rows, 17):
        ws4.cell(row=i, column=1, value=name)
        ws4.cell(row=i, column=2, value=lift)
    ws4["A16"] = "Experiment"
    ws4["B16"] = "Lift_%"
    _style_header(ws4, 16)

    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Relative Lift (%)"
    chart2.y_axis.title = "Lift %"
    n_lift = 16 + len(lift_rows)
    data2 = Reference(ws4, min_col=2, min_row=16, max_row=n_lift)
    cats2 = Reference(ws4, min_col=1, min_row=17, max_row=n_lift)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.height = 15
    chart2.width = 22
    ws4.add_chart(chart2, "D15")

    # ── Sheet 5: Executive Summary ────────────────────────────────────────
    ws5 = wb.create_sheet("Executive_Summary", 4)
    ws5["A1"] = "A/B Test Executive Summary"
    ws5["A1"].font = Font(bold=True, size=16)
    ws5["A3"] = "Total Experiments"
    ws5["B3"] = len(summary_df)
    ws5["A4"] = "Statistically Significant"
    ws5["B4"] = int((summary_df["P_Value"] < 0.05).sum())
    ws5["A5"] = "Recommended to Launch"
    ws5["B5"] = int((summary_df["Recommendation"] == "LAUNCH").sum())
    ws5["A6"] = "Projected Revenue Impact (Monthly)"
    ws5["B6"] = summary_df["Revenue_Impact_Monthly"].sum()
    ws5["B6"].number_format = '"$"#,##0.00'

    ws5["A8"] = "Experiment"
    ws5["B8"] = "P-Value"
    ws5["C8"] = "Significant"
    ws5["D8"] = "Recommendation"
    ws5["E8"] = "Revenue Impact"
    _style_header(ws5, 8)

    for r, row in summary_df.iterrows():
        rr = r + 9
        ws5.cell(row=rr, column=1, value=row["Experiment"])
        ws5.cell(row=rr, column=2, value=row["P_Value"])
        ws5.cell(row=rr, column=2).number_format = "0.00E+00"
        ws5.cell(row=rr, column=3, value=row["Significant"])
        ws5.cell(row=rr, column=4, value=row["Recommendation"])
        ws5.cell(row=rr, column=5, value=row["Revenue_Impact_Monthly"])
        ws5.cell(row=rr, column=5).number_format = '"$"#,##0.00'

    # Conditional formatting: green if p < 0.05, red otherwise
    n_summary = len(summary_df) + 8
    ws5.conditional_formatting.add(
        f"B9:B{n_summary}",
        FormulaRule(formula=[f"B9<0.05"], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")),
    )
    ws5.conditional_formatting.add(
        f"B9:B{n_summary}",
        FormulaRule(formula=[f"B9>=0.05"], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
    )

    ws5.column_dimensions["A"].width = 40
    for c in "BCDE":
        ws5.column_dimensions[c].width = 16

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    out = build_workbook()
    print(f"Workbook saved: {out}")
